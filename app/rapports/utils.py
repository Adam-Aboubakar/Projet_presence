import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models import Presence, Session, Personne, Configuration


# ============================================================
# UTILITAIRES COMMUNS
# ============================================================

def get_config():
    return Configuration.query.first()

def nom_etab():
    config = get_config()
    return config.nom_etablissement if config and config.nom_etablissement else "Établissement"

def style_excel_header():
    fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center")
    return fill, font, alignment

def bordure_excel():
    side = Side(style='thin', color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def dt_debut(d):
    return datetime.combine(d, datetime.min.time())

def dt_fin(d):
    return datetime.combine(d, datetime.max.time())

FILL_PRESENT = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FILL_RETARD  = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
FILL_ABSENT  = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

def fill_for_statut(statut):
    return {'present': FILL_PRESENT, 'retard': FILL_RETARD}.get(statut, FILL_ABSENT)

def couleur_statut(statut):
    return {
        'present': colors.HexColor('#27AE60'),
        'retard':  colors.HexColor('#F39C12'),
        'absent':  colors.HexColor('#E74C3C'),
    }.get(statut, colors.black)


# ============================================================
# PDF — Rapport complet étudiant
# ============================================================

def generer_pdf_etudiant(personne_id, date_debut, date_fin):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    personne = Personne.query.get(personne_id)
    if not personne:
        return None

    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle('titre', parent=styles['Title'],
                                 fontSize=16, textColor=colors.HexColor('#1A5276'),
                                 alignment=TA_CENTER)
    sous_titre_style = ParagraphStyle('sous_titre', parent=styles['Normal'],
                                      fontSize=11, textColor=colors.HexColor('#555555'),
                                      alignment=TA_CENTER)

    elements = []
    elements.append(Paragraph(nom_etab(), titre_style))
    elements.append(Paragraph("Rapport de Présence", sous_titre_style))
    elements.append(Spacer(1, 0.5*cm))

    infos_data = [
        ['Nom complet',        f'{personne.prenom} {personne.nom}'],
        ['Identifiant',        personne.identifiant or '-'],
        ['Filière / Département', personne.departement or '-'],
        ['Groupe',             personne.groupe_ou_site or '-'],
        ['Période',            f'{date_debut.strftime("%d/%m/%Y")} → {date_fin.strftime("%d/%m/%Y")}'],
        ['Généré le',          datetime.now().strftime("%d/%m/%Y à %H:%M")],
    ]
    infos_table = Table(infos_data, colWidths=[5*cm, 12*cm])
    infos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D6EAF8')),
        ('FONTNAME',   (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 10),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('PADDING',    (0, 0), (-1, -1), 6),
    ]))
    elements.append(infos_table)
    elements.append(Spacer(1, 0.5*cm))

    presences = Presence.query.join(Session).filter(
        Presence.personne_id == personne_id,
        Session.heure_debut >= dt_debut(date_debut),
        Session.heure_debut <= dt_fin(date_fin)
    ).order_by(Session.heure_debut).all()

    elements.append(Paragraph("Détail des Présences", ParagraphStyle(
        'section', parent=styles['Heading2'],
        textColor=colors.HexColor('#1A5276'), fontSize=12
    )))
    elements.append(Spacer(1, 0.3*cm))

    tableau_data = [['Session', 'Date', 'Heure', 'Statut', 'Modifié par']]
    total = present = retard = absent_just = absent_injust = 0

    for p in presences:
        session = Session.query.get(p.session_id)
        if not session:
            continue
        total += 1
        if p.statut == 'present':
            present += 1
        elif p.statut == 'retard':
            retard += 1
        elif p.statut == 'absent':
            if p.justification_absence:
                absent_just += 1
            else:
                absent_injust += 1

        tableau_data.append([
            session.nom[:30],
            session.heure_debut.strftime('%d/%m/%Y'),
            p.horodatage.strftime('%H:%M') if p.horodatage else '-',
            p.statut.upper(),
            p.modifie_par or '-'
        ])

    if len(tableau_data) > 1:
        t = Table(tableau_data, colWidths=[6*cm, 3*cm, 2*cm, 3*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, 0), colors.HexColor('#1A5276')),
            ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
            ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, -1), 9),
            ('GRID',         (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
            ('ALIGN',        (0, 0), (-1, -1), 'CENTER'),
            ('PADDING',      (0, 0), (-1, -1), 5),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("Aucune présence enregistrée sur cette période.",
                                  styles['Normal']))

    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph("Résumé", ParagraphStyle(
        'section2', parent=styles['Heading2'],
        textColor=colors.HexColor('#1A5276'), fontSize=12
    )))

    taux = round((present + retard) / total * 100, 1) if total > 0 else 0
    stats_data = [
        ['Total sessions',     str(total),       '100%'],
        ['Présent',            str(present),     f'{round(present/total*100,1) if total else 0}%'],
        ['Retard',             str(retard),      f'{round(retard/total*100,1)  if total else 0}%'],
        ['Absent justifié',    str(absent_just),   f'{round(absent_just/total*100,1)   if total else 0}%'],
        ['Absent injustifié',  str(absent_injust), f'{round(absent_injust/total*100,1) if total else 0}%'],
        ['Taux de présence',   f'{taux}%',       ''],
    ]
    stats_table = Table(stats_data, colWidths=[7*cm, 4*cm, 6*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D6EAF8')),
        ('FONTNAME',   (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 10),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('PADDING',    (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D5F5E3')),
        ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(stats_table)

    if absent_injust >= 3:
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(
            f"Attention : {absent_injust} absences injustifiees - Verifier le reglement de l'etablissement.",
            ParagraphStyle('alerte', parent=styles['Normal'],
                           textColor=colors.HexColor('#E74C3C'),
                           fontSize=10, fontName='Helvetica-Bold')
        ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ============================================================
# EXCEL — Session unique
# ============================================================

def generer_excel_session(session_id):
    """Génère un Excel pour une seule session — liste des présences."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Présences"

    fill_header, font_header, align_header = style_excel_header()
    bordure = bordure_excel()

    seance = Session.query.get(session_id)
    if not seance:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # Titre
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{nom_etab()} — {seance.nom}"
    ws['A1'].font = Font(bold=True, size=14, color="1A5276")
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:F2')
    ws['A2'] = (f"Date : {seance.heure_debut.strftime('%d/%m/%Y %H:%M')}"
                f" → {seance.heure_fin.strftime('%H:%M')}"
                f"  |  Lieu : {seance.lieu or '-'}")
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(size=11, color="555555")

    # En-têtes
    headers = ['Nom', 'Prénom', 'Identifiant', 'Statut', 'Heure pointage', 'Méthode']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_header
        cell.border = bordure

    # Données
    presences = Presence.query.filter_by(session_id=session_id)\
        .order_by(Presence.horodatage).all()

    for row_idx, p in enumerate(presences, 5):
        personne = Personne.query.get(p.personne_id)
        if not personne:
            continue

        donnees = [
            personne.nom,
            personne.prenom,
            personne.identifiant,
            p.statut.upper(),
            p.horodatage.strftime('%H:%M:%S') if p.horodatage else '-',
            p.methode_validation or '-'
        ]

        for col, val in enumerate(donnees, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill_for_statut(p.statut)
            cell.border = bordure
            cell.alignment = Alignment(horizontal="center")

    # Largeurs
    for i, w in enumerate([20, 15, 15, 12, 16, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================
# EXCEL — Résumé groupe
# ============================================================

def generer_excel_resume(groupe, date_debut, date_fin):
    """Génère un Excel résumé — une ligne par étudiant."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résumé"

    fill_header, font_header, align_header = style_excel_header()
    bordure = bordure_excel()

    ws.merge_cells('A1:H1')
    ws['A1'] = f"{nom_etab()} — Rapport Résumé Groupe {groupe}"
    ws['A1'].font = Font(bold=True, size=14, color="1A5276")
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Période : {date_debut.strftime('%d/%m/%Y')} → {date_fin.strftime('%d/%m/%Y')}"
    ws['A2'].alignment = Alignment(horizontal="center")

    headers = ['Nom', 'Prénom', 'Identifiant', 'Présent', 'Retard',
               'Absent Justifié', 'Absent Injustifié', 'Taux de Présence']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_header
        cell.border = bordure

    personnes = Personne.query.filter_by(
        groupe_ou_site=groupe, est_actif=True
    ).order_by(Personne.nom).all()

    for row_idx, personne in enumerate(personnes, 5):
        presences = Presence.query.join(Session).filter(
            Presence.personne_id == personne.id,
            Session.heure_debut >= dt_debut(date_debut),
            Session.heure_debut <= dt_fin(date_fin)
        ).all()

        total       = len(presences)
        present     = sum(1 for p in presences if p.statut == 'present')
        retard      = sum(1 for p in presences if p.statut == 'retard')
        absent_just = sum(1 for p in presences if p.statut == 'absent' and p.justification_absence)
        absent_injust = sum(1 for p in presences if p.statut == 'absent' and not p.justification_absence)
        taux        = round((present + retard) / total * 100, 1) if total > 0 else 0

        donnees = [personne.nom, personne.prenom, personne.identifiant,
                   present, retard, absent_just, absent_injust, f'{taux}%']

        bg = "FFFFFF" if row_idx % 2 == 0 else "F9F9F9"
        fill_row = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        for col, val in enumerate(donnees, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill_row
            cell.border = bordure
            cell.alignment = Alignment(horizontal="center")
            if col == 8:
                if taux >= 80:
                    cell.font = Font(color="27AE60", bold=True)
                elif taux >= 60:
                    cell.font = Font(color="F39C12", bold=True)
                else:
                    cell.font = Font(color="E74C3C", bold=True)

    for i, largeur in enumerate([20, 15, 15, 10, 10, 18, 20, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = largeur

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================
# EXCEL — Détail groupe (format liste professionnel)
# Lignes = une présence par étudiant par session
# Trié par étudiant → date
# ============================================================

def generer_excel_detail(groupe, date_debut, date_fin):
    """Génère un Excel détail professionnel — une ligne par présence."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Détail Présences"

    fill_header, font_header, align_header = style_excel_header()
    bordure = bordure_excel()

    # ── Titre ──
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{nom_etab()} — Rapport Détail Groupe {groupe}"
    ws['A1'].font = Font(bold=True, size=14, color="1A5276")
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Période : {date_debut.strftime('%d/%m/%Y')} → {date_fin.strftime('%d/%m/%Y')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(size=11, color="555555")

    # ── En-têtes ──
    headers = ['Nom', 'Prénom', 'Identifiant', 'Session', 'Date', 'Heure', 'Statut', 'Méthode']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_header
        cell.border = bordure

    # ── Données ──
    personnes = Personne.query.filter_by(
        groupe_ou_site=groupe, est_actif=True
    ).order_by(Personne.nom, Personne.prenom).all()

    row_idx = 5
    for personne in personnes:
        presences = Presence.query.join(Session).filter(
            Presence.personne_id == personne.id,
            Session.heure_debut >= dt_debut(date_debut),
            Session.heure_debut <= dt_fin(date_fin)
        ).order_by(Session.heure_debut).all()

        if not presences:
            # Ligne vide pour montrer l'étudiant sans présence
            donnees = [personne.nom, personne.prenom, personne.identifiant,
                       '—', '—', '—', 'AUCUNE DONNÉE', '—']
            for col, val in enumerate(donnees, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = bordure
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(color="999999", italic=True)
            row_idx += 1
            continue

        for p in presences:
            session = Session.query.get(p.session_id)
            if not session:
                continue

            donnees = [
                personne.nom,
                personne.prenom,
                personne.identifiant,
                session.nom,
                session.heure_debut.strftime('%d/%m/%Y'),
                p.horodatage.strftime('%H:%M') if p.horodatage else '—',
                p.statut.upper(),
                p.methode_validation or '—'
            ]

            for col, val in enumerate(donnees, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = bordure
                cell.alignment = Alignment(horizontal="center")
                # Colorer la ligne selon statut
                if col == 7:
                    cell.fill = fill_for_statut(p.statut)
                    cell.font = Font(bold=True)
                else:
                    bg = "FFFFFF" if row_idx % 2 == 0 else "F9F9F9"
                    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

            row_idx += 1

        # Ligne séparatrice entre étudiants (fond gris léger)
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col, value='')
            cell.fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")
        row_idx += 1

    # ── Largeurs colonnes ──
    for i, w in enumerate([18, 15, 15, 28, 12, 10, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Figer la ligne d'en-tête ──
    ws.freeze_panes = 'A5'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================
# EXCEL — Complet (résumé + détail)
# ============================================================

def _copier_feuille(ws_src, wb_final, titre):
    ws_copy = wb_final.create_sheet(title=titre)
    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_copy.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font      = cell.font.copy()
                new_cell.fill      = cell.fill.copy()
                new_cell.border    = cell.border.copy()
                new_cell.alignment = cell.alignment.copy()
    for col, dim in ws_src.column_dimensions.items():
        ws_copy.column_dimensions[col].width = dim.width
    for merge in ws_src.merged_cells.ranges:
        ws_copy.merge_cells(str(merge))


def generer_excel_complet(groupe, date_debut, date_fin):
    """Génère un Excel avec 2 onglets : Résumé + Détail."""
    wb_resume = openpyxl.load_workbook(generer_excel_resume(groupe, date_debut, date_fin))
    wb_detail = openpyxl.load_workbook(generer_excel_detail(groupe, date_debut, date_fin))

    wb_final = openpyxl.Workbook()
    wb_final.remove(wb_final.active)

    for ws_src in wb_resume.worksheets:
        _copier_feuille(ws_src, wb_final, "Résumé")
    for ws_src in wb_detail.worksheets:
        _copier_feuille(ws_src, wb_final, "Détail")

    buffer = io.BytesIO()
    wb_final.save(buffer)
    buffer.seek(0)
    return buffer
